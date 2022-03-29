import openpyxl

wb= openpyxl.load_workbook('test.xlsx')
sheetnames = wb.get_sheet_names()
print(sheetnames)