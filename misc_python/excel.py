import openpyxl

wb = openpyxl.load_workbook('example.xlsx')
print(wb.get_sheet_names())
print(type(wb))

sheet = wb.get_sheet_by_name('Sheet1')
c = (sheet['A1']).value
print(c)
d = sheet['B1'].value
print(d)

sheet2 = wb.get_sheet_by_name('Next')
e = (sheet2['C1']).value
print(e)

sheet3 = wb.get_sheet_by_name('Sheet3')
f = (sheet3['D1']).value
print(f)

print(sheet.cell(row=1, column=2).value)
for i in range(1, 8,):
    print(i, sheet.cell(row=i, column=2).value)


wb1 = openpyxl.Workbook()
wb1.get_sheet_names()
sheet = wb1.get_active_sheet()
print(sheet.title)
sheet['A1'] = 'Hello World'
wb1.save('example2.xlsx')

x = sheet['A1'].value
print(x)
